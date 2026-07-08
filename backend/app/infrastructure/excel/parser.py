from collections.abc import Sequence
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import cast
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from app.application.excel_contract import (
    IMPORT_COLUMNS,
    IMPORT_WORKSHEET_NAME,
    PREDECESSOR_SEPARATOR,
    REQUIRED_IMPORT_COLUMNS,
    ParsedTaskRow,
    normalize_header,
    normalize_task_name,
)
from app.application.excel_errors import ExcelValidationError, ExcelValidationIssue, issue

WorkbookCell = Cell | MergedCell


def parse_import_workbook(content: bytes) -> tuple[ParsedTaskRow, ...]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
    except (InvalidFileException, OSError, ValueError, KeyError, BadZipFile) as error:
        raise ExcelValidationError(
            [
                issue(
                    "malformed_workbook",
                    "The uploaded file is not a readable .xlsx workbook.",
                )
            ]
        ) from error

    worksheet = cast(
        Worksheet,
        workbook[IMPORT_WORKSHEET_NAME]
        if IMPORT_WORKSHEET_NAME in workbook.sheetnames
        else workbook.active,
    )
    rows = tuple(worksheet.iter_rows())
    if not rows:
        raise ExcelValidationError(
            [
                issue(
                    "empty_workbook",
                    "The workbook does not contain a header row.",
                    worksheet=worksheet.title,
                )
            ]
        )

    header_map, header_issues = _build_header_map(rows[0], worksheet)
    if header_issues:
        raise ExcelValidationError(header_issues)

    parsed_rows: list[ParsedTaskRow] = []
    validation_issues: list[ExcelValidationIssue] = []
    task_names: dict[str, int] = {}

    for row in rows[1:]:
        if _is_empty_row(row, header_map):
            continue

        parsed = _parse_row(row, header_map, worksheet.title)
        validation_issues.extend(parsed.issues)
        if parsed.row is not None:
            normalized_task_name = normalize_task_name(parsed.row.task)
            previous_row = task_names.get(normalized_task_name)
            if previous_row is not None:
                validation_issues.append(
                    issue(
                        "duplicate_task",
                        "Task names must be unique in the imported workbook.",
                        worksheet=worksheet.title,
                        row=parsed.row.row_number,
                        column="task",
                    )
                )
                validation_issues.append(
                    issue(
                        "duplicate_task",
                        "Task names must be unique in the imported workbook.",
                        worksheet=worksheet.title,
                        row=previous_row,
                        column="task",
                    )
                )
            else:
                task_names[normalized_task_name] = parsed.row.row_number
            parsed_rows.append(parsed.row)

    if not parsed_rows and not validation_issues:
        validation_issues.append(
            issue(
                "empty_workbook",
                "The workbook does not contain any task rows.",
                worksheet=worksheet.title,
            )
        )

    validation_issues.extend(_validate_predecessors(parsed_rows, task_names, worksheet.title))

    if validation_issues:
        raise ExcelValidationError(validation_issues)

    return tuple(parsed_rows)


def validate_xlsx_upload(
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    max_size_bytes: int,
) -> None:
    issues: list[ExcelValidationIssue] = []

    if Path(filename).suffix.casefold() != ".xlsx":
        issues.append(issue("unsupported_file_type", "Only .xlsx files are supported."))
    if content_type and content_type not in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    }:
        issues.append(
            issue("unsupported_content_type", "The uploaded file must be an .xlsx workbook.")
        )
    if len(content) > max_size_bytes:
        issues.append(
            issue(
                "file_too_large",
                f"The uploaded workbook must be {max_size_bytes} bytes or smaller.",
            )
        )
    if not content:
        issues.append(issue("empty_file", "The uploaded file is empty."))

    if issues:
        raise ExcelValidationError(issues)


class _ParsedRowResult:
    def __init__(self, row: ParsedTaskRow | None, issues: list[ExcelValidationIssue]) -> None:
        self.row = row
        self.issues = issues


def _build_header_map(
    header_row: Sequence[WorkbookCell],
    worksheet: Worksheet,
) -> tuple[dict[str, int], list[ExcelValidationIssue]]:
    header_map: dict[str, int] = {}
    issues: list[ExcelValidationIssue] = []

    for index, cell in enumerate(header_row):
        if _contains_formula(cell):
            issues.append(
                issue(
                    "formula_not_allowed",
                    "Formulas are not allowed in the import header.",
                    worksheet=worksheet.title,
                    row=cell.row,
                    column=str(cell.value),
                )
            )
            continue

        header = normalize_header(cell.value)
        if not header:
            continue
        if header in IMPORT_COLUMNS and header not in header_map:
            header_map[header] = index

    for column in REQUIRED_IMPORT_COLUMNS:
        if column not in header_map:
            issues.append(
                issue(
                    "missing_required_column",
                    f"Missing required column: {column}.",
                    worksheet=worksheet.title,
                    row=1,
                    column=column,
                )
            )

    return header_map, issues


def _parse_row(
    row: Sequence[WorkbookCell],
    header_map: dict[str, int],
    worksheet_name: str,
) -> _ParsedRowResult:
    row_number = row[0].row or 0
    issues: list[ExcelValidationIssue] = []

    for column in IMPORT_COLUMNS:
        cell = _cell_for_column(row, header_map, column)
        if cell is not None and _contains_formula(cell):
            issues.append(
                issue(
                    "formula_not_allowed",
                    "Formulas are not allowed in imported values.",
                    worksheet=worksheet_name,
                    row=row_number,
                    column=column,
                )
            )

    task_value = _text_cell(row, header_map, "task")
    if not task_value:
        issues.append(
            issue(
                "missing_task",
                "Task must be a non-empty text value.",
                worksheet=worksheet_name,
                row=row_number,
                column="task",
            )
        )

    duration = _duration_cell(row, header_map, worksheet_name, row_number, issues)
    description = _text_cell(row, header_map, "description") or ""
    assignee = _text_cell(row, header_map, "assignee") or None
    predecessors = _predecessor_cell(row, header_map, worksheet_name, row_number, issues)

    if issues or task_value is None or duration is None:
        return _ParsedRowResult(None, issues)

    return _ParsedRowResult(
        ParsedTaskRow(
            row_number=row_number,
            task=task_value,
            description=description,
            assignee=assignee,
            duration_days=duration,
            predecessor_names=predecessors,
        ),
        [],
    )


def _validate_predecessors(
    rows: list[ParsedTaskRow],
    task_names: dict[str, int],
    worksheet_name: str,
) -> list[ExcelValidationIssue]:
    issues: list[ExcelValidationIssue] = []

    for row in rows:
        normalized_task_name = normalize_task_name(row.task)
        seen_predecessors: set[str] = set()

        for predecessor_name in row.predecessor_names:
            normalized_predecessor = normalize_task_name(predecessor_name)
            if normalized_predecessor == normalized_task_name:
                issues.append(
                    issue(
                        "self_dependency",
                        "A task cannot depend on itself.",
                        worksheet=worksheet_name,
                        row=row.row_number,
                        column="predecessors",
                    )
                )
            if normalized_predecessor not in task_names:
                issues.append(
                    issue(
                        "unknown_predecessor",
                        f"Unknown predecessor task: {predecessor_name}.",
                        worksheet=worksheet_name,
                        row=row.row_number,
                        column="predecessors",
                    )
                )
            if normalized_predecessor in seen_predecessors:
                issues.append(
                    issue(
                        "duplicate_predecessor",
                        f"Duplicate predecessor task: {predecessor_name}.",
                        worksheet=worksheet_name,
                        row=row.row_number,
                        column="predecessors",
                    )
                )
            seen_predecessors.add(normalized_predecessor)

    return issues


def _cell_for_column(
    row: Sequence[WorkbookCell],
    header_map: dict[str, int],
    column: str,
) -> WorkbookCell | None:
    index = header_map.get(column)
    if index is None or index >= len(row):
        return None
    return row[index]


def _text_cell(row: Sequence[WorkbookCell], header_map: dict[str, int], column: str) -> str | None:
    cell = _cell_for_column(row, header_map, column)
    if cell is None or cell.value is None:
        return None
    value = str(cell.value).strip()
    return value or None


def _duration_cell(
    row: Sequence[WorkbookCell],
    header_map: dict[str, int],
    worksheet_name: str,
    row_number: int,
    issues: list[ExcelValidationIssue],
) -> int | None:
    cell = _cell_for_column(row, header_map, "duration")
    value = cell.value if cell is not None else None

    if value is None or str(value).strip() == "":
        issues.append(
            issue(
                "missing_duration",
                "Duration must be a positive integer.",
                worksheet=worksheet_name,
                row=row_number,
                column="duration",
            )
        )
        return None

    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        decimal_value = Decimal("-1.5")

    if decimal_value != decimal_value.to_integral_value() or decimal_value <= 0:
        issues.append(
            issue(
                "invalid_duration",
                "Duration must be a positive integer.",
                worksheet=worksheet_name,
                row=row_number,
                column="duration",
            )
        )
        return None

    return int(decimal_value)


def _predecessor_cell(
    row: Sequence[WorkbookCell],
    header_map: dict[str, int],
    worksheet_name: str,
    row_number: int,
    issues: list[ExcelValidationIssue],
) -> tuple[str, ...]:
    value = _text_cell(row, header_map, "predecessors")
    if value is None:
        return ()

    names = tuple(name.strip() for name in value.split(PREDECESSOR_SEPARATOR) if name.strip())
    if not names and value.strip():
        issues.append(
            issue(
                "invalid_predecessors",
                "Predecessors must be separated by semicolons.",
                worksheet=worksheet_name,
                row=row_number,
                column="predecessors",
            )
        )
    return names


def _is_empty_row(row: Sequence[WorkbookCell], header_map: dict[str, int]) -> bool:
    return all(_text_cell(row, header_map, column) is None for column in IMPORT_COLUMNS)


def _contains_formula(cell: WorkbookCell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")
