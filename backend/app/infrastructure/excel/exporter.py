from io import BytesIO
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.application.excel_contract import (
    EXPORT_COLUMNS,
    IMPORT_WORKSHEET_NAME,
    PREDECESSOR_SEPARATOR,
    XLSX_CONTENT_TYPE,
)
from app.domain.entities import Plan, Task


def export_plan_to_xlsx(plan: Plan) -> bytes:
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = IMPORT_WORKSHEET_NAME

    _write_header(worksheet)

    task_by_id = {task.id: task for task in plan.tasks}
    predecessor_names_by_task = _predecessor_names_by_task(plan, task_by_id)

    sorted_tasks = sorted(plan.tasks, key=lambda item: (item.position, item.id))
    for row_index, task in enumerate(sorted_tasks, start=2):
        worksheet.cell(row=row_index, column=1, value=task.id)
        worksheet.cell(row=row_index, column=2, value=task.name)
        worksheet.cell(row=row_index, column=3, value=task.description)
        worksheet.cell(row=row_index, column=4, value=task.assignee)
        worksheet.cell(row=row_index, column=5, value=task.duration_days)
        worksheet.cell(
            row=row_index,
            column=6,
            value=PREDECESSOR_SEPARATOR.join(predecessor_names_by_task.get(task.id, ())),
        )
        start_cell = worksheet.cell(row=row_index, column=7, value=task.start_date)
        end_cell = worksheet.cell(row=row_index, column=8, value=task.end_date)
        start_cell.number_format = "yyyy-mm-dd"
        end_cell.number_format = "yyyy-mm-dd"
        worksheet.cell(row=row_index, column=3).alignment = Alignment(wrap_text=True)

    _format_sheet(worksheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_filename(plan: Plan) -> str:
    normalized = "".join(
        character if character.isalnum() else "-"
        for character in plan.name.lower()
    )
    collapsed = "-".join(part for part in normalized.split("-") if part)
    return f"{collapsed or plan.id}.xlsx"


def _write_header(worksheet: Worksheet) -> None:
    fill = PatternFill("solid", fgColor="E8EEF7")
    for column_index, header in enumerate(EXPORT_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill


def _format_sheet(worksheet: Worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = {
        "A": 24,
        "B": 28,
        "C": 44,
        "D": 18,
        "E": 12,
        "F": 32,
        "G": 14,
        "H": 14,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _predecessor_names_by_task(
    plan: Plan,
    task_by_id: dict[str, Task],
) -> dict[str, tuple[str, ...]]:
    names: dict[str, list[str]] = {task.id: [] for task in plan.tasks}
    task_order = {task.id: (task.position, task.id) for task in plan.tasks}

    for dependency in sorted(
        plan.dependencies,
        key=lambda item: (
            task_order.get(item.successor_task_id, (0, "")),
            item.predecessor_task_id,
        ),
    ):
        predecessor = task_by_id.get(dependency.predecessor_task_id)
        if predecessor is None:
            continue
        names.setdefault(dependency.successor_task_id, []).append(predecessor.name)

    return {task_id: tuple(task_names) for task_id, task_names in names.items()}


__all__ = ["XLSX_CONTENT_TYPE", "export_filename", "export_plan_to_xlsx"]
